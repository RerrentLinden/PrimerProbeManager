import io
from collections import Counter
from datetime import date, datetime

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.primer import Primer
from app.models.primer_tube import PrimerTube
from app.schemas.import_data import (
    PrimerPreviewRow,
    TubePreviewRow,
    ImportPreviewResponse,
    ImportConfirmResponse,
)
from app.services.primer_metrics import calculate_gc_percent, count_bases
from app.services import tube_lifecycle_log_service

PRIMER_HEADERS = [
    "名称", "序列(5'→3')", "5'修饰", "3'修饰",
    "MW", "ug/OD", "nmol/OD", "GC%", "Tm", "纯化方式",
]
TUBE_HEADERS = ["引物名称", "批号", "定容日期", "产量(μL)", "项目"]


def generate_template() -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "引物信息"
    ws1.append(PRIMER_HEADERS)

    ws2 = wb.create_sheet("管信息")
    ws2.append(TUBE_HEADERS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def parse_preview(
    session: AsyncSession, file: UploadFile,
) -> ImportPreviewResponse:
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)

    primer_rows = _parse_primer_sheet(wb)
    tube_rows = _parse_tube_sheet(wb)

    primer_previews = await _preview_primers(session, primer_rows)
    tube_previews = await _preview_tubes(session, tube_rows, primer_rows)

    return ImportPreviewResponse(
        primers=primer_previews,
        tubes=tube_previews,
        primer_create_count=sum(1 for p in primer_previews if p.action == "create"),
        primer_update_count=sum(1 for p in primer_previews if p.action == "update"),
        tube_create_count=sum(1 for t in tube_previews if t.action == "create"),
        tube_update_count=sum(1 for t in tube_previews if t.action == "update"),
        error_count=sum(1 for t in tube_previews if t.action == "error"),
    )


async def confirm_import(
    session: AsyncSession, file: UploadFile,
) -> ImportConfirmResponse:
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)

    primer_rows = _parse_primer_sheet(wb)
    tube_rows = _parse_tube_sheet(wb)

    p_created, p_updated = await _import_primers(session, primer_rows)
    t_created, t_updated = await _import_tubes(session, tube_rows)

    await session.commit()
    return ImportConfirmResponse(
        primers_created=p_created,
        primers_updated=p_updated,
        tubes_created=t_created,
        tubes_updated=t_updated,
    )


def _parse_primer_sheet(wb: Workbook) -> list[dict]:
    ws = wb["引物信息"] if "引物信息" in wb.sheetnames else wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        rows.append({
            "name": str(row[0]).strip(),
            "sequence": str(row[1]).strip(),
            "modification_5prime": _str_or_none(row, 2),
            "modification_3prime": _str_or_none(row, 3),
            "mw": _float_or_none(row, 4),
            "ug_per_od": _float_or_none(row, 5),
            "nmol_per_od": _float_or_none(row, 6),
            "gc_percent": _normalize_gc(_float_or_none(row, 7)),
            "tm": _float_or_none(row, 8),
            "purification_method": _str_or_none(row, 9),
        })
    return rows


def _parse_tube_sheet(wb: Workbook) -> list[dict]:
    ws = wb["管信息"] if "管信息" in wb.sheetnames else wb.worksheets[1]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        rows.append({
            "primer_name": str(row[0]).strip(),
            "batch_number": str(row[1]).strip(),
            "dissolution_date": _parse_date(row, 2),
            "initial_volume_ul": _float_or_none(row, 3) or 0.0,
            "project": _str_or_none(row, 4),
        })
    return rows


async def _preview_primers(
    session: AsyncSession, rows: list[dict],
) -> list[PrimerPreviewRow]:
    results = []
    for r in rows:
        existing = await _find_primer_by_identity(session, r["name"], r["mw"])
        if existing and existing.sequence == r["sequence"]:
            action, msg = "update", "已存在，将更新其他字段"
        elif existing:
            action, msg = "conflict", f"同名同 MW 但序列不同(库中: {existing.sequence[:20]}...)"
        else:
            action, msg = "create", "新增"
        results.append(PrimerPreviewRow(**r, action=action, message=msg))
    return results


async def _preview_tubes(
    session: AsyncSession, rows: list[dict], primer_rows: list[dict],
) -> list[TubePreviewRow]:
    results = []
    known_names = {r["name"] for r in primer_rows}
    known_name_counts = Counter(r["name"] for r in primer_rows)
    for r in rows:
        db_matches = await _find_primers_by_name(session, r["primer_name"])
        if not db_matches and r["primer_name"] not in known_names:
            action, msg = "error", "引探不存在且未在 Sheet1 中定义"
        elif len(db_matches) > 1 or known_name_counts[r["primer_name"]] > 1:
            action, msg = "error", "同名引探存在多个 MW，当前模板无法匹配"
        else:
            primer_id = db_matches[0].id if db_matches else None
            existing_tube = await _find_tube(
                session, primer_id, r["batch_number"],
            )
            if existing_tube:
                action, msg = "update", "同批号已存在，将更新产量"
            else:
                action, msg = "create", "新增"
        results.append(TubePreviewRow(
            **r,
            dissolution_date=str(r["dissolution_date"]) if r["dissolution_date"] else None,
            action=action,
            message=msg,
        ))
    return results


async def _import_primers(
    session: AsyncSession, rows: list[dict],
) -> tuple[int, int]:
    created = updated = 0
    for r in rows:
        existing = await _find_primer_by_identity(session, r["name"], r["mw"])
        if existing:
            if existing.sequence == r["sequence"]:
                _apply_primer_fields(existing, r)
                updated += 1
        else:
            primer = Primer(**_build_primer_payload(r))
            session.add(primer)
            created += 1
    await session.flush()
    return created, updated


async def _import_tubes(
    session: AsyncSession, rows: list[dict],
) -> tuple[int, int]:
    created = updated = 0
    for r in rows:
        primers = await _find_primers_by_name(session, r["primer_name"])
        if not primers:
            continue
        if len(primers) > 1:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=(
                    f"导入失败：{r['primer_name']} 存在多个同名引探，"
                    "当前模板无法区分 MW"
                ),
            )
        primer = primers[0]
        existing = await _find_tube(session, primer.id, r["batch_number"])
        if existing:
            existing.initial_volume_ul = r["initial_volume_ul"]
            existing.remaining_volume_ul = r["initial_volume_ul"]
            if r["dissolution_date"]:
                existing.dissolution_date = r["dissolution_date"]
            if r["project"]:
                existing.project = r["project"]
            updated += 1
        else:
            tube = PrimerTube(
                primer_id=primer.id,
                batch_number=r["batch_number"],
                dissolution_date=r["dissolution_date"],
                initial_volume_ul=r["initial_volume_ul"],
                remaining_volume_ul=r["initial_volume_ul"],
                project=r["project"],
            )
            session.add(tube)
            await session.flush()
            tube_lifecycle_log_service.stage_created_log(
                session, tube=tube, primer_name=primer.name, primer_type=primer.type,
            )
            created += 1
    await session.flush()
    return created, updated


def _apply_primer_fields(primer: Primer, data: dict) -> None:
    skip = {"name", "sequence"}
    for key, val in data.items():
        if key not in skip and val is not None:
            setattr(primer, key, val)
    primer.base_count = count_bases(primer.sequence)
    primer.gc_percent = calculate_gc_percent(primer.sequence)


async def _find_primer_by_identity(
    session: AsyncSession, name: str, mw: float | None,
) -> Primer | None:
    query = select(Primer).where(Primer.name == name)
    if mw is None:
        query = query.where(Primer.mw.is_(None))
    else:
        query = query.where(Primer.mw == mw)
    return (
        await session.execute(query)
    ).scalar_one_or_none()


async def _find_tube(
    session: AsyncSession, primer_id: int | None, batch: str,
) -> PrimerTube | None:
    if primer_id is None:
        return None
    query = (
        select(PrimerTube)
        .where(PrimerTube.primer_id == primer_id, PrimerTube.batch_number == batch)
    )
    return (await session.execute(query)).scalar_one_or_none()


async def _find_primers_by_name(
    session: AsyncSession, name: str,
) -> list[Primer]:
    return list((
        await session.execute(
            select(Primer).where(Primer.name == name).order_by(Primer.id.desc())
        )
    ).scalars().all())


def _str_or_none(row: tuple, idx: int) -> str | None:
    if idx >= len(row) or row[idx] is None:
        return None
    val = str(row[idx]).strip()
    return val if val else None


def _float_or_none(row: tuple, idx: int) -> float | None:
    if idx >= len(row) or row[idx] is None:
        return None
    try:
        return float(row[idx])
    except (ValueError, TypeError):
        return None


def _parse_date(row: tuple, idx: int) -> date | None:
    if idx >= len(row) or row[idx] is None:
        return None
    val = row[idx]
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _normalize_gc(val: float | None) -> float | None:
    if val is None:
        return None
    # If > 1, assume percentage, convert to 0~1
    if val > 1:
        return val / 100.0
    return val


def _build_primer_payload(row: dict) -> dict:
    payload = {**row}
    payload["base_count"] = count_bases(payload["sequence"])
    payload["gc_percent"] = calculate_gc_percent(payload["sequence"])
    return payload
