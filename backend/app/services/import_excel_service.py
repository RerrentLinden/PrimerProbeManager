import io
from dataclasses import dataclass
from datetime import date, datetime

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook

PRIMER_SHEET_NAME = "引探信息"
TUBE_SHEET_NAME = "分管信息"

PRIMER_HEADERS = [
    "名称*",
    "序列(5'→3')*",
    "5'修饰",
    "3'修饰",
    "MW*",
    "ug/OD",
    "nmol/OD",
    "GC%",
    "Tm",
    "纯化方式",
    "项目",
]
TUBE_HEADERS = [
    "引探名称*",
    "引探MW*",
    "批号*",
    "分管编号",
    "定容日期*",
    "产量(μL)*",
    "冻存盒",
    "孔位",
]
PROJECT_SEPARATORS = ("，", ",", "；", ";", "、", "\n")


@dataclass(frozen=True)
class PrimerSheetRow:
    row_number: int
    name: str
    sequence: str
    modification_5prime: str | None
    modification_3prime: str | None
    mw: float
    ug_per_od: float | None
    nmol_per_od: float | None
    gc_percent: float | None
    tm: float | None
    purification_method: str | None
    project_names: tuple[str, ...]

    @property
    def import_key(self) -> str:
        return f"primer:{self.row_number}"


@dataclass(frozen=True)
class TubeSheetRow:
    row_number: int
    primer_name: str
    primer_mw: float
    batch_number: str
    tube_number: str | None
    dissolution_date: date
    initial_volume_ul: float
    box_name: str | None = None
    well_position: str | None = None

    @property
    def import_key(self) -> str:
        return f"tube:{self.row_number}"


def generate_template() -> bytes:
    workbook = Workbook()
    primer_sheet = workbook.active
    primer_sheet.title = PRIMER_SHEET_NAME
    primer_sheet.append(PRIMER_HEADERS)
    tube_sheet = workbook.create_sheet(TUBE_SHEET_NAME)
    tube_sheet.append(TUBE_HEADERS)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_workbook(content: bytes) -> tuple[list[PrimerSheetRow], list[TubeSheetRow]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    _ensure_sheet(workbook, PRIMER_SHEET_NAME)
    _ensure_sheet(workbook, TUBE_SHEET_NAME)
    return _parse_primers(workbook), _parse_tubes(workbook)


def _ensure_sheet(workbook: Workbook, name: str) -> None:
    if name not in workbook.sheetnames:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail=f"导入模板缺少工作表: {name}",
        )


def _parse_primers(workbook: Workbook) -> list[PrimerSheetRow]:
    sheet = workbook[PRIMER_SHEET_NAME]
    rows: list[PrimerSheetRow] = []
    for row_number, values in _iter_rows(sheet):
        if _is_blank(values):
            continue
        name = _require_string(values, 0, row_number, PRIMER_SHEET_NAME)
        sequence = _require_string(values, 1, row_number, PRIMER_SHEET_NAME)
        rows.append(
            PrimerSheetRow(
                row_number=row_number,
                name=name,
                sequence=sequence,
                modification_5prime=_optional_string(values, 2),
                modification_3prime=_optional_string(values, 3),
                mw=_require_float(values, 4, row_number, PRIMER_SHEET_NAME),
                ug_per_od=_optional_float(values, 5, row_number, PRIMER_SHEET_NAME),
                nmol_per_od=_optional_float(values, 6, row_number, PRIMER_SHEET_NAME),
                gc_percent=_normalize_gc(
                    _optional_float(values, 7, row_number, PRIMER_SHEET_NAME)
                ),
                tm=_optional_float(values, 8, row_number, PRIMER_SHEET_NAME),
                purification_method=_optional_string(values, 9),
                project_names=_optional_project_names(values, 10),
            )
        )
    return rows


def _parse_tubes(workbook: Workbook) -> list[TubeSheetRow]:
    sheet = workbook[TUBE_SHEET_NAME]
    rows: list[TubeSheetRow] = []
    for row_number, values in _iter_rows(sheet):
        if _is_blank(values):
            continue
        box_name = _optional_string(values, 6)
        well_position = _optional_string(values, 7)
        if box_name and not well_position:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                detail=f"{TUBE_SHEET_NAME} 第 {row_number} 行填写了冻存盒但缺少孔位",
            )
        if well_position and not box_name:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                detail=f"{TUBE_SHEET_NAME} 第 {row_number} 行填写了孔位但缺少冻存盒",
            )
        if well_position:
            _validate_well_format(well_position, row_number)
        rows.append(
            TubeSheetRow(
                row_number=row_number,
                primer_name=_require_string(values, 0, row_number, TUBE_SHEET_NAME),
                primer_mw=_require_float(values, 1, row_number, TUBE_SHEET_NAME),
                batch_number=_require_string(values, 2, row_number, TUBE_SHEET_NAME),
                tube_number=_optional_string(values, 3),
                dissolution_date=_require_date(values, 4, row_number, TUBE_SHEET_NAME),
                initial_volume_ul=_require_float(values, 5, row_number, TUBE_SHEET_NAME),
                box_name=box_name,
                well_position=well_position.upper() if well_position else None,
            )
        )
    return rows


def _iter_rows(sheet):
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        yield row_number, values


def _is_blank(values: tuple) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _require_string(values: tuple, index: int, row: int, sheet: str) -> str:
    value = _optional_string(values, index)
    if value is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet} 第 {row} 行第 {index + 1} 列不能为空",
        )
    return value


def _optional_string(values: tuple, index: int) -> str | None:
    if index >= len(values) or values[index] is None:
        return None
    text = str(values[index]).strip()
    return text or None


def _optional_project_names(values: tuple, index: int) -> tuple[str, ...]:
    raw_value = _optional_string(values, index)
    if raw_value is None:
        return ()
    normalized = raw_value
    for separator in PROJECT_SEPARATORS:
        normalized = normalized.replace(separator, ",")
    names: list[str] = []
    seen: set[str] = set()
    for item in normalized.split(","):
        name = item.strip()
        if not name or name in seen:
            continue
        names.append(name)
        seen.add(name)
    return tuple(names)


def _require_float(values: tuple, index: int, row: int, sheet: str) -> float:
    value = _optional_float(values, index, row, sheet)
    if value is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet} 第 {row} 行第 {index + 1} 列需要数值",
        )
    return value


def _require_date(values: tuple, index: int, row: int, sheet: str) -> date:
    value = _optional_date(values, index, row, sheet)
    if value is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet} 第 {row} 行第 {index + 1} 列不能为空",
        )
    return value


def _optional_float(values: tuple, index: int, row: int, sheet: str) -> float | None:
    if index >= len(values) or values[index] is None:
        return None
    try:
        return float(values[index])
    except (TypeError, ValueError) as exc:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail=f"{sheet} 第 {row} 行第 {index + 1} 列不是有效数字",
        ) from exc


def _optional_date(values: tuple, index: int, row: int, sheet: str) -> date | None:
    if index >= len(values) or values[index] is None:
        return None
    value = values[index]
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if isinstance(value, (int, float)):
        text = str(int(value))
    for fmt in ("%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise HTTPException(
        status.HTTP_400_BAD_REQUEST,
        detail=f"{sheet} 第 {row} 行第 {index + 1} 列日期格式应为 YYYY-MM-DD 或 YYYYMMDD",
    )


def parse_well_position(well: str) -> tuple[int, int]:
    row_char = well[0].upper()
    row = ord(row_char) - ord("A")
    col = int(well[1:]) - 1
    return row, col


def _validate_well_format(well: str, row_number: int) -> None:
    if len(well) < 2 or not well[0].isalpha() or not well[1:].isdigit():
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail=f"{TUBE_SHEET_NAME} 第 {row_number} 行孔位格式无效(应为 A1、B2 等)",
        )


def _normalize_gc(value: float | None) -> float | None:
    if value is None:
        return None
    if value > 1:
        return value / 100.0
    return value
