import io

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.box_position import BoxPosition
from app.models.freezer_box import FreezerBox
from app.models.primer import Primer
from app.models.primer_tube import PrimerTube
from app.models.project import Project
from app.models.project_gene import ProjectGene
from app.models.project_primer import ProjectPrimer


async def export_primers_xlsx(session: AsyncSession) -> bytes:
    primers = await _load_primers(session)
    project_names = await _load_primer_project_names(session)
    workbook = Workbook()
    primer_sheet = workbook.active
    primer_sheet.title = "引探信息"
    primer_sheet.append(
        [
            "引探ID",
            "名称",
            "类型",
            "序列",
            "5'修饰",
            "3'修饰",
            "MW",
            "Tm",
            "GC%",
            "纯化方式",
            "项目",
            "创建时间",
            "更新时间",
        ]
    )
    tube_sheet = workbook.create_sheet("分管信息")
    tube_sheet.append(
        [
            "分管ID",
            "引探ID",
            "引探名称",
            "批号",
            "分管编号",
            "状态",
            "定容日期",
            "初始体积(uL)",
            "剩余体积(uL)",
            "项目",
            "盒子ID",
            "盒子名称",
            "孔位行",
            "孔位列",
            "存储位置",
            "存储温度",
            "创建时间",
            "更新时间",
        ]
    )
    for primer in primers:
        primer_sheet.append(
            [
                primer.id,
                primer.name,
                primer.type,
                primer.sequence,
                primer.modification_5prime,
                primer.modification_3prime,
                primer.mw,
                primer.tm,
                primer.gc_percent,
                primer.purification_method,
                ", ".join(project_names.get(primer.id, [])),
                primer.created_at.isoformat(),
                primer.updated_at.isoformat(),
            ]
        )
        for tube in sorted(primer.tubes, key=lambda item: item.id):
            position = tube.position
            box = position.box if position else None
            tube_sheet.append(
                [
                    tube.id,
                    primer.id,
                    primer.name,
                    tube.batch_number,
                    tube.tube_number,
                    tube.status,
                    tube.dissolution_date.isoformat() if tube.dissolution_date else None,
                    tube.initial_volume_ul,
                    tube.remaining_volume_ul,
                    tube.project,
                    position.box_id if position else None,
                    box.name if box else None,
                    _display_row(position.row) if position else None,
                    _display_col(position.col) if position else None,
                    box.storage_location if box else None,
                    box.storage_temperature if box else None,
                    tube.created_at.isoformat(),
                    tube.updated_at.isoformat(),
                ]
            )
    return _workbook_bytes(workbook)


async def export_boxes_xlsx(session: AsyncSession) -> bytes:
    boxes = await _load_boxes(session)
    workbook = Workbook()
    box_sheet = workbook.active
    box_sheet.title = "冻存盒"
    box_sheet.append(
        [
            "盒子ID",
            "名称",
            "行数",
            "列数",
            "已占用孔位数",
            "存储位置",
            "存储温度",
            "创建时间",
            "更新时间",
        ]
    )
    position_sheet = workbook.create_sheet("盒位明细")
    position_sheet.append(
        [
            "盒子ID",
            "盒子名称",
            "孔位行",
            "孔位列",
            "分管ID",
            "引探ID",
            "引探名称",
            "类型",
            "批号",
            "分管编号",
            "状态",
            "剩余体积(uL)",
            "初始体积(uL)",
        ]
    )
    for box in boxes:
        box_sheet.append(
            [
                box.id,
                box.name,
                box.rows,
                box.cols,
                len(box.positions),
                box.storage_location,
                box.storage_temperature,
                box.created_at.isoformat(),
                box.updated_at.isoformat(),
            ]
        )
        for position in sorted(box.positions, key=lambda item: (item.row, item.col)):
            tube = position.tube
            primer = tube.primer if tube else None
            position_sheet.append(
                [
                    box.id,
                    box.name,
                    _display_row(position.row),
                    _display_col(position.col),
                    tube.id if tube else None,
                    primer.id if primer else None,
                    primer.name if primer else None,
                    primer.type if primer else None,
                    tube.batch_number if tube else None,
                    tube.tube_number if tube else None,
                    tube.status if tube else None,
                    tube.remaining_volume_ul if tube else None,
                    tube.initial_volume_ul if tube else None,
                ]
            )
    return _workbook_bytes(workbook)


async def export_projects_xlsx(session: AsyncSession) -> bytes:
    projects = await _load_projects(session)
    workbook = Workbook()
    project_sheet = workbook.active
    project_sheet.title = "项目信息"
    project_sheet.append(["项目ID", "名称", "描述", "引探数", "基因数", "创建时间", "更新时间"])
    primer_sheet = workbook.create_sheet("项目引探")
    primer_sheet.append(
        [
            "项目ID",
            "项目名称",
            "引探ID",
            "引探名称",
            "类型",
            "序列",
            "MW",
            "Tm",
            "GC%",
            "当前剩余总体积(uL)",
        ]
    )
    gene_sheet = workbook.create_sheet("项目基因面板")
    gene_sheet.append(["项目ID", "项目名称", "基因ID", "基因名", "管号", "荧光通道", "排序"])
    tube_sheet = workbook.create_sheet("项目分管存储")
    tube_sheet.append(
        [
            "项目ID",
            "项目名称",
            "引探ID",
            "引探名称",
            "分管ID",
            "批号",
            "分管编号",
            "状态",
            "剩余体积(uL)",
            "盒子名称",
            "孔位行",
            "孔位列",
            "存储位置",
            "存储温度",
        ]
    )
    for project in projects:
        project_sheet.append(
            [
                project.id,
                project.name,
                project.description,
                len(project.primer_links),
                len(project.genes),
                project.created_at.isoformat(),
                project.updated_at.isoformat(),
            ]
        )
        for link in project.primer_links:
            primer = link.primer
            if primer is None:
                continue
            total_remaining_volume = _active_tube_remaining_volume(primer.tubes)
            primer_sheet.append(
                [
                    project.id,
                    project.name,
                    primer.id,
                    primer.name,
                    primer.type,
                    primer.sequence,
                    primer.mw,
                    primer.tm,
                    primer.gc_percent,
                    total_remaining_volume,
                ]
            )
            for tube in sorted(primer.tubes, key=lambda item: item.id):
                position = tube.position
                box = position.box if position else None
                tube_sheet.append(
                    [
                        project.id,
                        project.name,
                        primer.id,
                        primer.name,
                        tube.id,
                        tube.batch_number,
                        tube.tube_number,
                        tube.status,
                        tube.remaining_volume_ul,
                        box.name if box else None,
                        _display_row(position.row) if position else None,
                        _display_col(position.col) if position else None,
                        box.storage_location if box else None,
                        box.storage_temperature if box else None,
                    ]
                )
        for gene in sorted(project.genes, key=lambda item: item.sort_order):
            gene_sheet.append(
                [project.id, project.name, gene.id, gene.gene_name, gene.tube_number, gene.fluorescence_channel, gene.sort_order]
            )
    return _workbook_bytes(workbook)


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _display_row(row: int) -> int:
    return row + 1


def _display_col(col: int) -> int:
    return col + 1


def _active_tube_remaining_volume(tubes: list[PrimerTube]) -> float:
    return sum(tube.remaining_volume_ul for tube in tubes if tube.status == "active")


async def _load_primers(session: AsyncSession) -> list[Primer]:
    query = select(Primer).options(
        selectinload(Primer.tubes).selectinload(PrimerTube.position).selectinload(BoxPosition.box)
    )
    return list((await session.execute(query.order_by(Primer.id))).scalars().all())


async def _load_primer_project_names(session: AsyncSession) -> dict[int, list[str]]:
    query = select(ProjectPrimer.primer_id, Project.name).join(Project, Project.id == ProjectPrimer.project_id)
    rows = (await session.execute(query)).all()
    result: dict[int, list[str]] = {}
    for primer_id, project_name in rows:
        result.setdefault(primer_id, []).append(project_name)
    return result


async def _load_boxes(session: AsyncSession) -> list[FreezerBox]:
    query = select(FreezerBox).options(
        selectinload(FreezerBox.positions).selectinload(BoxPosition.tube).selectinload(PrimerTube.primer)
    )
    return list((await session.execute(query.order_by(FreezerBox.id))).scalars().all())


async def _load_projects(session: AsyncSession) -> list[Project]:
    query = select(Project).options(
        selectinload(Project.primer_links)
        .selectinload(ProjectPrimer.primer)
        .selectinload(Primer.tubes)
        .selectinload(PrimerTube.position)
        .selectinload(BoxPosition.box),
        selectinload(Project.genes),
    )
    return list((await session.execute(query.order_by(Project.id))).scalars().all())
